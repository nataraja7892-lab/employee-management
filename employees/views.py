from django.shortcuts import render,redirect
from django.contrib.auth import authenticate, login,logout
from django.contrib import messages
from .models import *
from django.shortcuts import get_object_or_404
from django.contrib.auth.decorators import login_required
import json
from django.db.models import Q
from django.db.models import Case, When, Value, IntegerField

def landing_page(request):
    return render(request,'landing_page.html')

def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username') 
        password = request.POST.get('password')
        
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('dashboard') 
        else:
            messages.error(request, "Invalid username or password")
            return redirect('landing_page')
    
    return redirect('landing_page')

def logout_view(request):
    logout(request)
    messages.success(request, "You have been logged out successfully.")
    return redirect('landing_page')

@login_required
def dashboard_view(request):
    total_employees = Employee.objects.count()
    teaching_staff = Employee.objects.filter(category='Teaching').count()
    non_teaching_staff = Employee.objects.filter(category='Non-Teaching').count()
    phd_count = Employee.objects.filter(qualifications__icontains='phd').count()

    bsc_count = Employee.objects.filter(department__name='BSC').count()
    msc_count = Employee.objects.filter(department__name='MSC').count()
    bba_count = Employee.objects.filter(department__name='BBA').count()
    bca_count = Employee.objects.filter(department__name='BCA').count()
    ba_count = Employee.objects.filter(department__name='BA').count()
    ma_count = Employee.objects.filter(department__name='MA').count()
    bcom_count = Employee.objects.filter(department__name='BCOM').count()
    mcom_count = Employee.objects.filter(department__name='MCOM').count()

    context = {
        'total_employees': total_employees,
        'teaching_staff': teaching_staff,
        'non_teaching_staff': non_teaching_staff,
        'phd_count': phd_count,
        'bsc_count': bsc_count,
        'msc_count': msc_count,
        'bba_count': bba_count,
        'mcom_count': mcom_count,
        'bca_count': bca_count,
        'ba_count': ba_count,
        'ma_count': ma_count,
        'bcom_count': bcom_count,
    }
    return render(request, 'dashboard.html', context)

@login_required
def register_employee(request):
    departments = Department.objects.all()
    designations = Designation.objects.all()

    if request.method == 'POST':
        # BASIC VALIDATION
        employee_id = request.POST.get('employee_id')
        if Employee.objects.filter(employee_id=employee_id).exists():
            messages.error(request, "Employee ID already exists.")
            return redirect('register_employee')

        # BASIC DETAILS
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        short_name = request.POST.get('short_name')
        photo = request.FILES.get('photo')
        employee_type = request.POST.get('employee_type')
        category = request.POST.get('category')

        # PROFESSIONAL DETAILS
        department = Department.objects.filter(id=request.POST.get('department')).first()
        designation = Designation.objects.filter(id=request.POST.get('designation')).first()
        qualifications = request.POST.get('qualifications')
        present_working_as = request.POST.get('present_working_as')
        date_of_join = request.POST.get('date_of_join')
        date_of_exit = request.POST.get('date_of_exit')
        previous_experience_years = request.POST.get('previous_experience_years')
        experience_certificate = request.FILES.get('experience_certificate')

        # PERSONAL DETAILS
        gender = request.POST.get('gender')
        blood = request.POST.get('bloodGroup')
        date_of_birth = request.POST.get('dateOfBirth')
        marital_status = request.POST.get('maritalStatus')
        father_name = request.POST.get('fatherName')
        mother_name = request.POST.get('motherName')

        # CONTACT DETAILS
        mobile = request.POST.get('mobile')
        email = request.POST.get('email')
        address = request.POST.get('address')
        emergency_contact = request.POST.get('emergencyContact')

        # CREATE EMPLOYEE
        employee = Employee.objects.create(
            employee_id=employee_id,
            first_name=first_name,
            last_name=last_name,
            short_name=short_name,
            photo=photo,
            employee_type=employee_type,
            category=category,
            department=department,
            designation=designation,
            qualifications=qualifications
        )

        # RELATED TABLES
        EmployeeProfessionalDetails.objects.create(
            employee=employee,
            present_working_as=present_working_as,
            date_of_join=date_of_join or None,
            date_of_exit=date_of_exit or None,
            previous_experience_years=previous_experience_years or 0,
            experience_certificate=experience_certificate
        )

        EmployeePersonalDetails.objects.create(
            employee=employee,
            gender=gender,
            blood_group=blood,
            date_of_birth=date_of_birth,
            marital_status=marital_status,
            father_name=father_name,
            mother_name=mother_name
        )

        EmployeeContactDetails.objects.create(
            employee=employee,
            mobile_number=mobile,
            email=email,
            permanent_address=address,
            contact_address=emergency_contact
        )

        EmployeeDocuments.objects.create(
            employee=employee,
            aadhar_number=request.POST.get('aadharNumber'),
            aadhar_certificate=request.FILES.get('aadharDocument'),
            pan_number=request.POST.get('panNumber'),
            pan_certificate=request.FILES.get('panDocument'),
            joining_letter=request.FILES.get('joiningLetter'),
            tenth_certificate=request.FILES.get('tenthCertificate'),
            puc_certificate=request.FILES.get('pucCertificate'),
            degree_certificate=request.FILES.get('degreeCertificate'),
            pg_certificate=request.FILES.get('pgCertificate'),
            phd_certificate=request.FILES.get('phdCertificate')
        )

        EmployeeBankDetails.objects.create(
            employee=employee,
            pf_account_number=request.POST.get('pfAccountNumber'),
            esi_number=request.POST.get('esiNumber'),
            uan_number=request.POST.get('uanNumber'),
            bank_name=request.POST.get('bankName'),
            branch_name=request.POST.get('branchName'),
            ifsc_code=request.POST.get('ifscCode'),
            account_number=request.POST.get('accountNumber'),
            bank_passbook=request.FILES.get('bankPassbook')
        )

        messages.success(request, "✅ Employee registered successfully!")
        return redirect('dashboard')

    return render(request, 'register_employee.html', {
        'departments': departments,
        'designations': designations,
    })

@login_required
def employee_list(request):
    employees = Employee.objects.all()
    context = {'employees': employees}
    return render(request, 'employee_list.html', context)

@login_required
def department_details_view(request, dept_name):
    # Get department safely
    department = get_object_or_404(Department, name=dept_name)

    # Get all employees in this department
    employees = Employee.objects.filter(department=department)

    # Calculate stats
    total_employees = employees.count()
    permanent_count = employees.filter(employee_type='Permanent').count()
    temporary_count = employees.filter(employee_type='Temporary').count()

    # Get gender counts from related EmployeePersonalDetails
    male_count = EmployeePersonalDetails.objects.filter(employee__in=employees, gender='Male').count()
    female_count = EmployeePersonalDetails.objects.filter(employee__in=employees, gender='Female').count()

    context = {
        'department': department,
        'employees': employees,
        'total_employees': total_employees,
        'permanent_count': permanent_count,
        'temporary_count': temporary_count,
        'male_count': male_count,
        'female_count': female_count,
    }
    return render(request, 'department_details.html', context)

def employee_profile_view(request, emp_id):
    employee = get_object_or_404(Employee, employee_id=emp_id)
    
    # Fetch related details
    personal = getattr(employee, 'employeepersonaldetails', None)
    professional = getattr(employee, 'employeeprofessionaldetails', None)
    contact = getattr(employee, 'employeecontactdetails', None)
    documents = getattr(employee, 'employeedocuments', None)
    bank = getattr(employee, 'employeebankdetails', None)

    context = {
        'employee': employee,
        'personal': personal,
        'professional': professional,
        'contact': contact,
        'documents': documents,
        'bank': bank,
    }
    return render(request, 'employee_profile.html', context)

@login_required
def edit_dashboard(request):
    departments=Department.objects.all()
    designations=Designation.objects.all()
    employees = Employee.objects.select_related('designation', 'department').all()

    employee_list = [
        {
            "id": emp.employee_id,
            "name": f"{emp.first_name} {emp.last_name}",
            "gender": emp.employeepersonaldetails.gender if emp.employeepersonaldetails else "",
            "designation": emp.designation.name if emp.designation else "",
            "department": emp.department.name if emp.department else "",
        }
        for emp in employees
    ]

    return render(request, 'edit_dashboard.html', {
        'departments': departments,
        'designations': designations,
        'employees_json': json.dumps(employee_list)
    })

@login_required
def edit_employee(request, emp_id):
    employee = get_object_or_404(Employee, employee_id=emp_id)
    departments = Department.objects.all()
    designations = Designation.objects.all()

    # Get or create related objects
    professional, _ = EmployeeProfessionalDetails.objects.get_or_create(employee=employee)
    personal, _ = EmployeePersonalDetails.objects.get_or_create(employee=employee)
    contact, _ = EmployeeContactDetails.objects.get_or_create(employee=employee)
    documents, _ = EmployeeDocuments.objects.get_or_create(employee=employee)
    bank, _ = EmployeeBankDetails.objects.get_or_create(employee=employee)

    if request.method == 'POST':
        # ---------------- BASIC DETAILS ----------------
        employee.first_name = request.POST.get('first_name')
        employee.last_name = request.POST.get('last_name')
        employee.short_name = request.POST.get('short_name')
        employee.employee_type = request.POST.get('employee_type')
        employee.category = request.POST.get('category')
        employee.department = Department.objects.filter(id=request.POST.get('department')).first()
        employee.designation = Designation.objects.filter(id=request.POST.get('designation')).first()
        employee.qualifications = request.POST.get('qualifications')

        # Photo replacement
        if 'photo' in request.FILES:
            if employee.photo:
                employee.photo.delete(save=False)
            employee.photo = request.FILES['photo']
        employee.save()

        # ---------------- PROFESSIONAL DETAILS ----------------
        professional.present_working_as = request.POST.get('present_working_as')
        professional.date_of_join = request.POST.get('date_of_join') or None
        professional.date_of_exit = request.POST.get('date_of_exit') or None
        professional.previous_experience_years = request.POST.get('previous_experience_years') or 0
        if 'experience_certificate' in request.FILES:
            if professional.experience_certificate:
                professional.experience_certificate.delete(save=False)
            professional.experience_certificate = request.FILES['experience_certificate']
        professional.save()

        # ---------------- PERSONAL DETAILS ----------------
        personal.gender = request.POST.get('gender')
        personal.blood_group = request.POST.get('bloodGroup')
        personal.date_of_birth = request.POST.get('dateOfBirth') or None
        personal.marital_status = request.POST.get('maritalStatus')
        personal.father_name = request.POST.get('fatherName')
        personal.mother_name = request.POST.get('motherName')
        personal.save()

        # ---------------- CONTACT DETAILS ----------------
        contact.mobile_number = request.POST.get('mobile')
        contact.email = request.POST.get('email')
        contact.permanent_address = request.POST.get('address')
        contact.contact_address = request.POST.get('emergencyContact')
        contact.save()

        # ---------------- DOCUMENTS ----------------
        documents.aadhar_number = request.POST.get('aadharNumber')
        if 'aadharDocument' in request.FILES:
            if documents.aadhar_certificate:
                documents.aadhar_certificate.delete(save=False)
            documents.aadhar_certificate = request.FILES['aadharDocument']
        documents.pan_number = request.POST.get('panNumber')
        if 'panDocument' in request.FILES:
            if documents.pan_certificate:
                documents.pan_certificate.delete(save=False)
            documents.pan_certificate = request.FILES['panDocument']
        if 'joiningLetter' in request.FILES:
            if documents.joining_letter:
                documents.joining_letter.delete(save=False)
            documents.joining_letter = request.FILES['joiningLetter']
        if 'tenthCertificate' in request.FILES:
            if documents.tenth_certificate:
                documents.tenth_certificate.delete(save=False)
            documents.tenth_certificate = request.FILES['tenthCertificate']
        if 'pucCertificate' in request.FILES:
            if documents.puc_certificate:
                documents.puc_certificate.delete(save=False)
            documents.puc_certificate = request.FILES['pucCertificate']
        if 'degreeCertificate' in request.FILES:
            if documents.degree_certificate:
                documents.degree_certificate.delete(save=False)
            documents.degree_certificate = request.FILES['degreeCertificate']
        if 'pgCertificate' in request.FILES:
            if documents.pg_certificate:
                documents.pg_certificate.delete(save=False)
            documents.pg_certificate = request.FILES['pgCertificate']
        if 'phdCertificate' in request.FILES:
            if documents.phd_certificate:
                documents.phd_certificate.delete(save=False)
            documents.phd_certificate = request.FILES['phdCertificate']
        documents.save()

        # ---------------- BANK DETAILS ----------------
        bank.pf_account_number = request.POST.get('pfAccountNumber')
        bank.esi_number = request.POST.get('esiNumber')
        bank.uan_number = request.POST.get('uanNumber')
        bank.bank_name = request.POST.get('bankName')
        bank.branch_name = request.POST.get('branchName')
        bank.ifsc_code = request.POST.get('ifscCode')
        bank.account_number = request.POST.get('accountNumber')
        if 'bankPassbook' in request.FILES:
            if bank.bank_passbook:
                bank.bank_passbook.delete(save=False)
            bank.bank_passbook = request.FILES['bankPassbook']
        bank.save()

        messages.success(request, "✅ Employee details updated successfully!")
        return redirect('employee_profile', emp_id=employee.employee_id)

    context = {
        'employee': employee,
        'departments': departments,
        'designations': designations,
        'professional': professional,
        'personal': personal,
        'contact': contact,
        'documents': documents,
        'bank': bank,
    }
    return render(request, 'edit_employee.html', context)

@login_required
def delete_employee(request, emp_id):
    employee = get_object_or_404(Employee, employee_id=emp_id)
    
    if request.method == "POST":
        employee.delete()
        messages.success(request, "Employee deleted successfully.")
        return redirect('dashboard')

    return render(request, 'confirm_delete.html', {'employee': employee})

@login_required
def search_employee(request):
    query = request.GET.get('q', '')
    employees = Employee.objects.filter(
        Q(first_name__icontains=query) | 
        Q(last_name__icontains=query) | 
        Q(employee_id__icontains=query) |
        Q(department__name__icontains=query) |
        Q(designation__name__icontains=query)
    ).distinct()
    return render(request, 'search_results.html', {'employees': employees, 'query': query})

from django.shortcuts import render
from django.utils.safestring import mark_safe
from django.http import HttpResponse, JsonResponse
from django.db.models import Count
import csv
import pandas as pd
import json
from datetime import datetime
import re
import io
from .models import Department, Employee, EmployeePersonalDetails, EmployeeProfessionalDetails

class EmployeeAnalytics:
    """Advanced analytics and data processing class"""
    
    CATEGORY_FIELD_MAPPING = {
        "basic": ["first_name", "last_name", "short_name", "photo", 
                 "employee_type", "category", "department", "designation", "qualifications"],
        "personal": ["gender", "date_of_birth", "blood_group", "marital_status", 
                    "father_name", "mother_name"],
        "professional": ["present_working_as", "date_of_join", "date_of_exit", 
                        "previous_experience_years", "experience_certificate"],
        "contact": ["mobile_number","alternate_mobile_number", "email", "permanent_address", "contact_address"],
        "bank": ["bank_name", "branch_name", "ifsc_code", "account_number", 
                "pf_account_number", "esi_number", "uan_number", "bank_passbook"],
        "documents": ["aadhar_number", "aadhar_certificate", "pan_number", "pan_certificate",
                     "joining_letter", "tenth_certificate", "puc_certificate", 
                     "degree_certificate", "pg_certificate", "phd_certificate"],
    }

    FIELD_RELATION_MAP = {
        "basic": "employee",
        "personal": "employeepersonaldetails", 
        "professional": "employeeprofessionaldetails",
        "contact": "employeecontactdetails",
        "bank": "employeebankdetails",
        "documents": "employeedocuments",
    }

    @classmethod
    def get_dashboard_stats(cls, department=None, staff_category=None):
        """
        Get dashboard stats.
        - department: str or None -> name of the department, or None for all
        - staff_category: 'Teaching', 'Non-Teaching', or None
        """

        qs = Employee.objects.all()

        # Filter by department (if a real department is selected)
        if department and department not in ["All", "Teaching", "Non-Teaching"]:
            qs = qs.filter(department__name__iexact=department)

        # Filter by staff category (Teaching / Non-Teaching)
        if staff_category in ["Teaching", "Non-Teaching"]:
            qs = qs.filter(designation__category__iexact=staff_category) # or, to include employees without designation, you'd need a more complex Q query

        total_employees = qs.count()
        total_departments = Department.objects.count()

        # Gender distribution
        gender_stats = EmployeePersonalDetails.objects.filter(employee__in=qs).values('gender').annotate(
            count=Count('gender')
        )

        # Department distribution
        dept_stats = qs.values('department__name').annotate(
            count=Count('employee_id')
        )

        # Teaching / Non-Teaching counts for stats cards
        teaching_count = qs.filter(designation__category="Teaching").count()
        non_teaching_count = qs.filter(designation__category="Non-Teaching").count()

        return {
            'total_employees': total_employees,
            'total_departments': total_departments,
            'gender_stats': list(gender_stats),
            'department_stats': list(dept_stats),
            'teaching_staff': teaching_count,
            'non_teaching_staff': non_teaching_count
        }

    @classmethod
    def get_field_value(cls, emp, category, field):
        """Advanced field value retrieval with formatting"""
        rel_attr = cls.FIELD_RELATION_MAP.get(category)
        
        if category == "basic":
            return cls._get_basic_field_value(emp, field)
        else:
            return cls._get_related_field_value(emp, rel_attr, field)

    @classmethod
    def _get_basic_field_value(cls, emp, field):
        """Get basic field values"""
        if field == "department":
            return emp.department.name if emp.department else "N/A"
        elif field == "designation":
            return emp.designation.name if emp.designation else "N/A"
        elif field == "photo":
            if emp.photo:
                return mark_safe(f'<img src="{emp.photo.url}" class="photo-thumbnail" alt="Employee Photo" data-bs-toggle="tooltip" title="View Photo">')
            return mark_safe('<i class="fas fa-user-circle text-muted fa-2x"></i>')
        else:
            val = getattr(emp, field, "N/A")
            return cls._format_value(val, field)

    @classmethod
    def _get_related_field_value(cls, emp, rel_attr, field):
        """Get related field values"""
        rel_obj = getattr(emp, rel_attr, None)
        if not rel_obj:
            return "N/A"
        
        val = getattr(rel_obj, field, None)
        if not val:
            return "N/A"
            
        return cls._format_value(val, field)

    @classmethod
    def _format_value(cls, val, field):
        """Format values based on field type"""
        if hasattr(val, 'url'):
            return cls._format_file_value(val, field)
        elif isinstance(val, datetime):
            return val.strftime('%Y-%m-%d %H:%M')
        else:
            return str(val) if val else "N/A"

    @classmethod
    def _format_file_value(cls, val, field):
        """Format file and image values"""
        file_name = str(val).split('/')[-1]
        if field.endswith('_certificate') or field in ['bank_passbook', 'joining_letter', 'experience_certificate']:
            return mark_safe(f'<a href="{val.url}" target="_blank" class="document-badge" data-bs-toggle="tooltip" title="Download {file_name}">📄 {file_name[:15]}...</a>')
        elif field == 'photo':
            return mark_safe(f'<img src="{val.url}" class="photo-thumbnail" alt="Photo" data-bs-toggle="tooltip" title="View Photo">')
        return val.url

    @classmethod
    def get_plain_text_value(cls, value):
        """Convert HTML value to plain text for export"""
        if hasattr(value, '__html__'):
            # Remove HTML tags
            text_value = re.sub('<[^<]+?>', '', str(value))
            return text_value.strip()
        return str(value) if value else "N/A"

def get_employee_export_data(selected_department, all_selected_fields):
    """Get employee data for export (CSV/Excel)"""
    employees_data = []

    if not all_selected_fields:
        return employees_data

    try:
        # Handle "All", "Teaching", "Non-Teaching", or specific department
        if selected_department == "All" or not selected_department:
            employees = Employee.objects.all()
        elif selected_department in ["Teaching", "Non-Teaching"]:
            employees = Employee.objects.filter(designation__category__iexact=selected_department)
        else:
            try:
                dept = Department.objects.get(name__iexact=selected_department)
                employees = Employee.objects.filter(department=dept)
            except Department.DoesNotExist:
                employees = Employee.objects.none()
                print(f"Department '{selected_department}' not found")
        
        for emp in employees:
            row_data = {
                "Employee ID": emp.employee_id,
                "Name": f"{emp.first_name} {emp.last_name}"
            }

            for category, fields in all_selected_fields.items():
                for field in fields:
                    column_name = f"{category.title()}: {field.replace('_', ' ').title()}"
                    value = EmployeeAnalytics.get_field_value(emp, category, field)
                    plain_value = EmployeeAnalytics.get_plain_text_value(value)
                    row_data[column_name] = plain_value

            employees_data.append(row_data)

    except Exception as e:
        print(f"Error in get_employee_export_data: {str(e)}")

    return employees_data

def export_csv(request):
    """Advanced CSV export with proper error handling"""
    try:
        selected_department = request.GET.get("department", "")
        all_selected_fields = {}
        
        # Collect selected fields
        for category in EmployeeAnalytics.CATEGORY_FIELD_MAPPING.keys():
            fields = request.GET.getlist(f"fields_{category}")
            if fields:
                all_selected_fields[category] = fields
        
        # Get export data
        employees_data = get_employee_export_data(selected_department, all_selected_fields)
        
        if not employees_data:
            return HttpResponse(
                "No data available for export. Please select a department and fields.",
                content_type='text/plain'
            )
        
        # Create CSV response
        response = HttpResponse(content_type='text/csv')
        filename = f"employee_data_{selected_department}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Create CSV writer
        writer = csv.writer(response)
        
        # Write headers
        if employees_data:
            headers = list(employees_data[0].keys())
            writer.writerow(headers)
            
            # Write data rows
            for employee in employees_data:
                row = [employee.get(header, 'N/A') for header in headers]
                writer.writerow(row)
        
        return response
        
    except Exception as e:
        return HttpResponse(
            f"Error generating CSV: {str(e)}",
            content_type='text/plain',
            status=500
        )

def export_excel(request):
    """Excel export with fallback options"""
    try:
        selected_department = request.GET.get("department", "")
        all_selected_fields = {}
        
        # Collect selected fields
        for category in EmployeeAnalytics.CATEGORY_FIELD_MAPPING.keys():
            fields = request.GET.getlist(f"fields_{category}")
            if fields:
                all_selected_fields[category] = fields
        
        # Get export data
        employees_data = get_employee_export_data(selected_department, all_selected_fields)
        
        if not employees_data:
            return HttpResponse(
                "No data available for export. Please select a department and fields.",
                content_type='text/plain'
            )
        
        # Try to use openpyxl for Excel export
        try:
            import openpyxl
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Employee Data"
            
            # Write headers
            if employees_data:
                headers = list(employees_data[0].keys())
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                
                # Write data
                for row, employee in enumerate(employees_data, 2):
                    for col, header in enumerate(headers, 1):
                        ws.cell(row=row, column=col, value=employee.get(header, 'N/A'))
                
                # Auto-adjust column widths
                for col in range(1, len(headers) + 1):
                    column_letter = openpyxl.utils.get_column_letter(col)
                    ws.column_dimensions[column_letter].width = 20
            
            # Create response
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            filename = f"employee_data_{selected_department}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            wb.save(response)
            return response
            
        except ImportError:
            # Fallback to pandas if openpyxl is not available
            try:
                df = pd.DataFrame(employees_data)
                
                response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                filename = f"employee_data_{selected_department}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                
                df.to_excel(response, index=False, engine='openpyxl')
                return response
                
            except ImportError:
                # Final fallback - create a simple Excel using CSV format
                return export_csv(request)
        
    except Exception as e:
        return HttpResponse(
            f"Error generating Excel file: {str(e)}",
            content_type='text/plain',
            status=500
        )

def export_pdf(request):
    """PDF export placeholder with better message"""
    return HttpResponse(
        "PDF export feature is coming soon! Currently we support CSV and Excel exports.",
        content_type='text/plain'
    )

@login_required
def employee_details(request):
    """Advanced employee details view with analytics"""
    from datetime import datetime
    
    departments = Department.objects.all()
    selected_department = request.GET.get("department", "")
    selected_category = request.GET.get("category", "")
    print(selected_category, selected_department)

    # ------------------------
    # Determine employees queryset
    # ------------------------
    if selected_department == "All" or not selected_department:
        employees_qs = Employee.objects.all()
    elif selected_department in ["Teaching", "Non-Teaching"]:
        employees_qs = Employee.objects.filter(designation__category__iexact=selected_department)
    else:
        try:
            dept = Department.objects.get(name__iexact=selected_department)
            employees_qs = Employee.objects.filter(department=dept)
        except Department.DoesNotExist:
            employees_qs = Employee.objects.none()
            message = f"Department '{selected_department}' not found"

    # Dashboard stats (no arguments)
    dashboard_stats = EmployeeAnalytics.get_dashboard_stats()

    # ------------------------
    # Selected fields
    # ------------------------

    # Add this block here
    field_label_overrides = {
        "pf_account_number": "PF Account Number",
        "esi_number": "ESI Number",
        "uan_number": "UAN Number"
    }

    all_selected_fields = {}
    all_selected_fields_labels = []

    for category in EmployeeAnalytics.CATEGORY_FIELD_MAPPING.keys():
        fields = request.GET.getlist(f"fields_{category}")
        if fields:
            all_selected_fields[category] = fields
            for field in fields:
                all_selected_fields_labels.append({
                    "category": category,
                    "name": field,
                    # Use override if exists, else default label
                    "label": field_label_overrides.get(field, field.replace('_', ' ').title())
                })

    # Default selected category
    if not selected_category and EmployeeAnalytics.CATEGORY_FIELD_MAPPING:
        selected_category = list(EmployeeAnalytics.CATEGORY_FIELD_MAPPING.keys())[0]

    # Prepare current category fields
    current_fields = EmployeeAnalytics.CATEGORY_FIELD_MAPPING.get(selected_category, [])
    current_selected = all_selected_fields.get(selected_category, [])

    available_fields = [
        {
            "name": f,
            "label": f.replace('_', ' ').title(),
            "display_label": field_label_overrides.get(f, f.replace('_', ' ').title()),
            "checked": f in current_selected
        }
        for f in current_fields
    ]

    # ------------------------
    # Build employees_data for display
    # ------------------------
    employees_data = []
    message = ""
    
    if employees_qs.exists() and all_selected_fields:
        for emp in employees_qs:
            row_data = {
                "employee_id": emp.employee_id,
                "name": f"{emp.first_name} {emp.last_name}",
                "values": []
            }
            for category, fields in all_selected_fields.items():
                for field in fields:
                    value = EmployeeAnalytics.get_field_value(emp, category, field)
                    row_data["values"].append(value)
            employees_data.append(row_data)
    else:
        if selected_department in ["All", "Teaching", "Non-Teaching"]:
            message = f"No employees found for {selected_department} staff"
        else:
            message = f"No employees found in {selected_department} department"

    # ------------------------
    # Prepare category metadata
    # ------------------------
    all_categories_data = {}
    for category in EmployeeAnalytics.CATEGORY_FIELD_MAPPING.keys():
        fields = EmployeeAnalytics.CATEGORY_FIELD_MAPPING[category]
        selected_count = len(all_selected_fields.get(category, []))
        all_categories_data[category] = {
            'selected_count': selected_count,
            'available_fields': [
                {"name": f, "label": f.replace('_', ' ').title(), "checked": f in all_selected_fields.get(category, [])}
                for f in fields
            ]
        }

    # ------------------------
    # Export params
    # ------------------------
    export_params = f"department={selected_department}"
    for category, fields in all_selected_fields.items():
        for field in fields:
            export_params += f"&fields_{category}={field}"

    # ------------------------
    # Context
    # ------------------------
    context = {
        "departments": departments,
        "selected_department": selected_department,
        "selected_category": selected_category,
        "available_fields": available_fields,
        "all_selected_fields_labels": all_selected_fields_labels,
        "employees": employees_data,
        "message": message,
        "categories": list(EmployeeAnalytics.CATEGORY_FIELD_MAPPING.keys()),
        "all_categories_data": all_categories_data,
        "total_selected_count": len(all_selected_fields_labels),
        "export_params": export_params,
        "total_employees": dashboard_stats['total_employees'],
        "current_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        **dashboard_stats
    }

    return render(request, "employee_details.html", context)


def analytics_view(request):
    # 1️⃣ Employees by Department
    department_data = (
        Department.objects
        .annotate(employee_count=Count('employee'))
        .values_list('name', 'employee_count')
    )
    department_labels = [d[0] for d in department_data]
    department_counts = [d[1] for d in department_data]

    # 2️⃣ Teaching vs Non-Teaching
    category_data = (
        Employee.objects
        .values('category')
        .annotate(count=Count('category'))
    )
    category_labels = [item['category'] for item in category_data]
    category_counts = [item['count'] for item in category_data]

    # 3️⃣ Permanent vs Temporary
    type_data = (
        Employee.objects
        .values('employee_type')
        .annotate(count=Count('employee_type'))
    )
    type_labels = [item['employee_type'] for item in type_data]
    type_counts = [item['count'] for item in type_data]

    context = {
        'department_labels': department_labels,
        'department_counts': department_counts,
        'category_labels': category_labels,
        'category_counts': category_counts,
        'type_labels': type_labels,
        'type_counts': type_counts,
    }
    return render(request, 'analytics.html', context)

def employee_list_all(request):
    employees = Employee.objects.all()

    employees = employees.annotate(
        staff_order=Case(
            When(designation__category='Teaching', then=Value(0)),
            When(designation__category='Non-Teaching', then=Value(1)),
            default=Value(2),
            output_field=IntegerField()
        )
    ).order_by('staff_order', 'employee_id')

    total_employees = employees.count()
    permanent_count = employees.filter(employee_type='Permanent').count()
    temporary_count = employees.filter(employee_type='Temporary').count()

    # Get gender counts from related EmployeePersonalDetails
    male_count = EmployeePersonalDetails.objects.filter(employee__in=employees, gender='Male').count()
    female_count = EmployeePersonalDetails.objects.filter(employee__in=employees, gender='Female').count()
    return render(request, 'employee_list.html', {
        'employees': employees,
        'page_title': 'All Employees',
        'description': 'Manage and view all employees across the institution.',
        'total_employees': total_employees,
        'permanent_count': permanent_count,
        'temporary_count': temporary_count,
        'male_count': male_count,
        'female_count': female_count,
    })

def employee_list_teaching(request):
    employees = Employee.objects.filter(category='Teaching')

    total_employees = employees.count()
    permanent_count = employees.filter(employee_type='Permanent').count()
    temporary_count = employees.filter(employee_type='Temporary').count()

    # Get gender counts from related EmployeePersonalDetails
    male_count = EmployeePersonalDetails.objects.filter(employee__in=employees, gender='Male').count()
    female_count = EmployeePersonalDetails.objects.filter(employee__in=employees, gender='Female').count()
    return render(request, 'employee_list.html', {
        'employees': employees,
        'page_title': 'Teaching Staff',
        'description': 'View and manage all teaching staff members.',
        'total_employees': total_employees,
        'permanent_count': permanent_count,
        'temporary_count': temporary_count,
        'male_count': male_count,
        'female_count': female_count,
    })

def employee_list_non_teaching(request):
    employees = Employee.objects.filter(category='Non-Teaching')

    total_employees = employees.count()
    permanent_count = employees.filter(employee_type='Permanent').count()
    temporary_count = employees.filter(employee_type='Temporary').count()

    # Get gender counts from related EmployeePersonalDetails
    male_count = EmployeePersonalDetails.objects.filter(employee__in=employees, gender='Male').count()
    female_count = EmployeePersonalDetails.objects.filter(employee__in=employees, gender='Female').count()
    return render(request, 'employee_list.html', {
        'employees': employees,
        'page_title': 'Non-Teaching Staff',
        'description': 'View and manage all non-teaching staff members.',
        'total_employees': total_employees,
        'permanent_count': permanent_count,
        'temporary_count': temporary_count,
        'male_count': male_count,
        'female_count': female_count,
    })

def employee_list_phd(request):
    employees = Employee.objects.filter(qualifications__icontains='phd')

    employees = employees.annotate(
        staff_order=Case(
            When(designation__category='Teaching', then=Value(0)),
            When(designation__category='Non-Teaching', then=Value(1)),
            default=Value(2),
            output_field=IntegerField()
        )
    ).order_by('staff_order', 'employee_id')

    total_employees = employees.count()
    permanent_count = employees.filter(employee_type='Permanent').count()
    temporary_count = employees.filter(employee_type='Temporary').count()

    # Get gender counts from related EmployeePersonalDetails
    male_count = EmployeePersonalDetails.objects.filter(employee__in=employees, gender='Male').count()
    female_count = EmployeePersonalDetails.objects.filter(employee__in=employees, gender='Female').count()
    return render(request, 'employee_list.html', {
        'employees': employees,
        'page_title': 'PhD Holders',
        'description': 'View employees who have completed PhD degrees.',
        'total_employees': total_employees,
        'permanent_count': permanent_count,
        'temporary_count': temporary_count,
        'male_count': male_count,
        'female_count': female_count,
    })
def he(request):
    return render(request,'form.html')
def gg(request):
    if request.method=='POST':
        name = request.POST.get('name')
        if not name:  # check if it's empty
           return redirect('hello')

        else:
            return HttpResponse(f"Hello, {name}!")
  
    else:
         return HttpResponse("Please submit the form.")
    
